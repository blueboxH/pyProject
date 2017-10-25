# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html

import openpyxl
from scrapy.exceptions import DropItem

class DuplicatesPipeline(object):

    def __init__(self):
        self.ids_seen = set()

    def process_item(self, item, spider):
        spider.log('执行去重')
        if item['id'] in self.ids_seen:
            raise DropItem("删除重复数据: %s" % item)
        else:
            self.ids_seen.add(item['id'])
            return item

class QianximapPipeline(object):

    def __init__(self, file_name):
        self.file_name = file_name

    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            file_name=crawler.settings.get('FULL_DATA')
        )

    def process_item(self, item, spider):
        spider.log('执行处理')
        try:
            ws = self.wb.get_sheet_by_name(item['date'])
        except KeyError:
            ws = self.wb.create_sheet(item['date'])
            ws.append(['排序标记', '出发地', '目的地', '人数', '汽车占比', '火车占比', '飞机占比'])

        ws.append([item['index'], item['start_adr'], item['end_adr'], item['count'], item['car'], item['train'], item['airplane']])

        return item

    def open_spider(self, spider):
        try:
            self.wb = openpyxl.load_workbook(self.file_name)
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()

    def close_spider(self, spider):
        try:
            self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))
        except KeyError:
            pass
        self.wb.save(self.file_name)


class filterPipeline(object):

    def __init__(self, file_name):
        self.file_name = file_name
        self.sheet_name = 'Sheet1'

    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            file_name=crawler.settings.get('FILTERED_CITY')
        )

    def process_item(self, item, spider):
        spider.log('执行筛选')
        if item['start_adr'] in self.cities or item['end_adr'] in self.cities:
            return item
        raise DropItem("删除不必要的数据: %s" % item)


    def open_spider(self, spider):
        self.wb = openpyxl.load_workbook(self.file_name)
        self.ws = self.wb.get_sheet_by_name(self.sheet_name)
        self.cities = [x.value for x in self.ws.get_cell_collection()]


    def close_spider(self, spider):
        self.wb.save(self.file_name)

class filteredQianximapPipeline(QianximapPipeline):
    @classmethod
    def from_crawler(cls, crawler):
        return cls(
            file_name=crawler.settings.get('FILTERED_DATA')
        )
